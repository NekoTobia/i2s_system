import pandas as pd
import os
from datasets import load_dataset 
from torch.utils.data import Dataset, DataLoader
from deep_translator import GoogleTranslator
import openpyxl
from PIL import Image
import urllib.request


class Create_datasets:
    def __init__(self,excel_name='default.xlsx',csv_name='default.csv',datasets_name='vizwiz',lang='zh_tw'):
        self.datasets_name = datasets_name
        self.excel_name = excel_name
        self.csv_name = csv_name
        self.lang = lang
        
    def creat_datasets(self,total):
        if self.datasets_name == 'vizwiz':
            split = 'train[:'+str(total)+']'
            dataset = load_dataset("Multimodal-Fatima/VizWiz_train",split=split)
            dataset = dataset.select_columns(['image','blip_caption_beam_5'])
            dataset = dataset.rename_column("blip_caption_beam_5", "text")
            path = 'image'
            if not os.path.isdir(path):
                os.mkdir(path)
            num=0
            for i in dataset:
                image_name = 'image/'+str(num) + '.jpg'
                num+=1
                i['image'].save(image_name,"jpeg")
                
        if self.datasets_name == 'mscoco':
            split = 'train[:'+str(total)+']'
            dataset = load_dataset("ChristophSchuhmann/MS_COCO_2017_URL_TEXT",split=split)
            dataset = dataset.rename_column("TEXT", "text")
            path = 'image'
            if not os.path.isdir(path):
                os.mkdir(path)
            num=0
            for i in dataset:
                image_name = 'image/'+str(num) + '.jpg'
                num+=1
                image = Image.open(urllib.request.urlopen(i['URL']))
                image.save(image_name,"jpeg")
        
        return dataset
        
    def creat_excel(self):
        wb = openpyxl.Workbook()
        s1 = wb.active
        s1.cell(1,1).value = 'image'
        s1.cell(1,2).value = 'text'
        wb.save(self.excel_name)
        return 'Save'

    def continue_translate_xlsx(self,dataset):
        #檢查基本資訊
        wb = openpyxl.load_workbook(self.excel_name)
        s1 = wb.active
        count = s1.max_row # 1 100
        check = 0

        #寫入
        if self.lang == 'zh_tw':
            for i in range(count,len(dataset)+1):
                j = i-1
                text = GoogleTranslator(source='en', target='zh-TW').translate(dataset['text'][j])
                image_name = 'image/'+str(j) + '.jpg'
                s1.cell(i+1,1).value = image_name
                s1.cell(i+1,2).value = text
                    
                check+=1
                if check%5 == 0:
                    print('已處理'+str(check)+'筆')
                    
                wb.save(self.excel_name)

        
        elif self.lang == 'en':
            for i in range(count,len(dataset)+1):
                j = i-1
                image_name = 'image/'+str(j) + '.jpg'
                s1.cell(i+1,1).value = image_name
                s1.cell(i+1,2).value = dataset['text'][j]
                    
                check+=1
                if check%5 == 0:
                    print('已處理'+str(check)+'筆')
                    
                wb.save(self.excel_name)
                    
        df = pd.DataFrame(pd.read_excel(self.excel_name))
        df.to_csv(self.csv_name)